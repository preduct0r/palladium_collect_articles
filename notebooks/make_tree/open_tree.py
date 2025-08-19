from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Optional


def _derive_default_output_path(xlsx_path: Path, num_rows: int) -> Path:
	return xlsx_path.with_name(f"{xlsx_path.stem}.head{num_rows}{xlsx_path.suffix}")


def _get_total_rows(xlsx_path: Path, sheet_name: Optional[str] = None) -> int:
	"""Get total number of rows in the Excel file."""
	try:
		import pandas as pd  # type: ignore
		df = pd.read_excel(
			xlsx_path,
			sheet_name=(sheet_name if sheet_name is not None else 0),
			engine=None,
		)
		if isinstance(df, dict):
			first_key = next(iter(df))
			df = df[first_key]
		return len(df)
	except ImportError:
		pass
	except Exception:
		pass
	
	# Fallback to openpyxl
	try:
		from openpyxl import load_workbook  # type: ignore
		wb = load_workbook(filename=xlsx_path, read_only=True)
		ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
		return ws.max_row - 1 if ws.max_row > 0 else 0  # -1 to exclude header
	except ImportError:
		return 0
	except Exception:
		return 0


def _format_file_size(size_bytes: int) -> str:
	"""Format file size in human-readable format."""
	for unit in ['B', 'KB', 'MB', 'GB']:
		if size_bytes < 1024.0:
			return f"{size_bytes:.1f} {unit}"
		size_bytes /= 1024.0
	return f"{size_bytes:.1f} TB"


def print_first_rows(xlsx_path: Path, num_rows: int = 10, sheet_name: Optional[str] = None, output_path: Optional[Path] = None) -> int:
	"""Print the header and first `num_rows` rows from an Excel file and optionally save them.

	Tries to use pandas if installed; falls back to openpyxl otherwise.
	Returns process exit code (0 for success, non-zero for errors).
	"""
	if not xlsx_path.exists():
		sys.stderr.write(f"Error: file not found: {xlsx_path}\n")
		return 2

	# Display file information
	total_rows = _get_total_rows(xlsx_path, sheet_name)
	print(f"File: {xlsx_path}")
	print(f"Total rows: {total_rows}")
	print(f"Showing first {num_rows} rows:")
	print("-" * 50)

	# Try pandas first for nicer formatting and easier export
	try:
		import pandas as pd  # type: ignore

		df = pd.read_excel(
			xlsx_path,
			sheet_name=(sheet_name if sheet_name is not None else 0),
			engine=None,
		)
		# If a dict is returned (when sheet_name=None), take the first sheet
		if isinstance(df, dict):
			first_key = next(iter(df))
			df = df[first_key]

		to_print = df.head(num_rows)
		# Print with header; human-friendly formatting
		print(to_print.to_string(index=False))

		# Save if requested
		if output_path is None:
			output_path = _derive_default_output_path(xlsx_path, num_rows)
		# Decide by extension
		ext = output_path.suffix.lower()
		if ext == ".csv":
			to_print.to_csv(output_path, index=False)
		elif ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
			to_print.to_excel(output_path, index=False)
		else:
			# Default to Excel if unknown
			to_print.to_excel(output_path.with_suffix(".xlsx"), index=False)
			sys.stderr.write(f"Unknown output extension '{ext}', wrote Excel instead: {output_path.with_suffix('.xlsx')}\n")

		return 0
	except ImportError:
		pass
	except Exception as exc:  # if pandas is present but cannot read, try openpyxl fallback
		sys.stderr.write(f"pandas failed to read Excel ({type(exc).__name__}: {exc}). Falling back to openpyxl...\n")

	# Fallback: openpyxl
	try:
		from openpyxl import load_workbook, Workbook  # type: ignore
	except ImportError:
		sys.stderr.write(
			"Neither pandas nor openpyxl is available. Install one of them, e.g.:\n"
			"  pip install pandas\n"
			"or\n"
			"  pip install openpyxl\n"
		)
		return 3

	wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
	ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

	rows_iter = ws.iter_rows(values_only=True)
	try:
		headers = next(rows_iter)
	except StopIteration:
		# Empty sheet
		return 0

	def stringify(values: tuple[object | None, ...]) -> str:
		return "\t".join("" if v is None else str(v) for v in values)

	print(stringify(tuple(headers)))
	collected: list[tuple[object | None, ...]] = []
	count = 0
	for row in rows_iter:
		print(stringify(tuple(row)))
		collected.append(tuple(row))
		count += 1
		if count >= num_rows:
			break

	# Save if requested (or default destination if not provided)
	if output_path is None:
		output_path = _derive_default_output_path(xlsx_path, num_rows)

	ext = output_path.suffix.lower()
	if ext == ".csv":
		import csv
		with open(output_path, "w", newline="", encoding="utf-8") as f:
			writer = csv.writer(f)
			writer.writerow(list(headers))
			for row in collected:
				writer.writerow(list(row))
	elif ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
		out_wb = Workbook(write_only=True)
		out_ws = out_wb.create_sheet()
		out_ws.append(list(headers))
		for row in collected:
			out_ws.append(list(row))
		out_wb.save(output_path)
	else:
		# Default to Excel if unknown
		fallback = output_path.with_suffix(".xlsx")
		out_wb = Workbook(write_only=True)
		out_ws = out_wb.create_sheet()
		out_ws.append(list(headers))
		for row in collected:
			out_ws.append(list(row))
		out_wb.save(fallback)
		sys.stderr.write(f"Unknown output extension '{ext}', wrote Excel instead: {fallback}\n")

	return 0


def main() -> int:
	script_dir = Path(__file__).parent
	default_xlsx = script_dir / "table_tree.xlsx"

	parser = argparse.ArgumentParser(description="Print the first N rows of an Excel file (with header) and optionally save them")
	parser.add_argument("-f", "--file", dest="file", type=Path, default=default_xlsx, help="Path to .xlsx file (default: table_tree.xlsx next to this script)")
	parser.add_argument("-s", "--sheet", dest="sheet", type=str, default=None, help="Sheet name to read (default: first sheet)")
	parser.add_argument("-n", dest="num_rows", type=int, default=10, help="Number of rows to print/save (default: 10)")
	parser.add_argument("-o", "--output", dest="output", type=Path, default=None, help="Output file path (e.g. out.xlsx or out.csv). Default: <input>.headN.xlsx")
	args = parser.parse_args()

	return print_first_rows(xlsx_path=args.file, num_rows=args.num_rows, sheet_name=args.sheet, output_path=args.output)


if __name__ == "__main__":
	sys.exit(main())
